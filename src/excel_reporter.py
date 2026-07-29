"""
excel_reporter.py
-----------------
Module de génération automatique du rapport de performance Excel.
Utilise openpyxl pour appliquer la charte graphique de Bouygues Telecom 
(Bleu Marine #003366, Bleu Ciel #0099CC, Orange #FF6600).
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Répertoire de sortie
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "outputs")

# --- CHARTE GRAPHIQUE BOUYGUES TELECOM ---
COLOR_NAVY = "003366"     # Bleu Marine (En-têtes principaux)
COLOR_SKY_BLUE = "0099CC" # Bleu Ciel (Titres & Sous-sections)
COLOR_ORANGE = "FF6600"   # Orange Bouygues (Accents & Highlights)
COLOR_BG_LIGHT = "F2F4F7"  # Gris léger (Lignes alternées)
COLOR_RED_ALERT = "FADBD8" # Rouge clair (Alerte dégradation)
COLOR_GREEN_OK = "D4EFDF"  # Vert clair (Statut conforme)

# Styles de bordure
THIN_BORDER = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)


def create_styled_excel_report(
    perf_summary: pd.DataFrame, 
    reco_summary: pd.DataFrame, 
    eco_metrics: dict, 
    output_filename: str = "Rapport_Performance_Reseau.xlsx"
) -> str:
    """
    Génère un rapport Excel multi-onglets stylisé aux couleurs de Bouygues Telecom.
    
    Args:
        perf_summary (pd.DataFrame): Données d'analyse par cellule.
        reco_summary (pd.DataFrame): Préconisations de dimensionnement ATOLL.
        eco_metrics (dict): Métriques d'impact environnemental (Sleep Mode).
        output_filename (str): Nom du fichier de sortie.
        
    Returns:
        str: Chemin d'accès au fichier Excel généré.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    excel_path = os.path.join(OUTPUT_DIR, output_filename)

    wb = Workbook()

    # -------------------------------------------------------------------------
    # ONGLET 1 : EXECUTIVE SUMMARY & RSE
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Titre Principal
    ws1.merge_cells("A1:E2")
    ws1["A1"] = "BTECHNOLOGIE / BOUYGUES TELECOM - COCKPIT NETWORK & RSE"
    ws1["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws1["A1"].fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Section 1 : KPIs Réseau Globaux
    ws1.merge_cells("A4:E4")
    ws1["A4"] = "📊 Synthèse Opérationnelle du Parc Radio (500 Sites)"
    ws1["A4"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws1["A4"].fill = PatternFill(start_color=COLOR_SKY_BLUE, end_color=COLOR_SKY_BLUE, fill_type="solid")

    kpi_cards = [
        ("Nombre Total de Cellules Analysées", len(perf_summary), "Cellules"),
        ("Nombre de Cellules Dégradées (Anomalies)", len(perf_summary[perf_summary['is_degraded']]), "Cellules"),
        ("Taux de Conformité Réseau", f"{(1 - len(perf_summary[perf_summary['is_degraded']])/len(perf_summary))*100:.1f}%", "%"),
        ("Trafic Total Écoulé (24h)", f"{perf_summary['total_traffic_gb'].sum():,.0f} Go", "Volume")
    ]

    for idx, (label, val, unit) in enumerate(kpi_cards, start=5):
        ws1[f"A{idx}"] = label
        ws1[f"D{idx}"] = val
        ws1[f"A{idx}"].font = Font(bold=True)
        ws1[f"D{idx}"].font = Font(bold=True, color=COLOR_ORANGE if "Dégradées" in label else "000000")
        ws1[f"A{idx}"].border = THIN_BORDER
        ws1[f"D{idx}"].border = THIN_BORDER

    # Section 2 : Impact Décarbonation & Énergie (MIMO Sleep Mode)
    ws1.merge_cells("A11:E11")
    ws1["A11"] = "🌱 Impact Économie Décarbonée (Optimisation Nocturne MIMO Sleep Mode)"
    ws1["A11"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws1["A11"].fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")

    rse_cards = [
        ("Cellules Optimisées en Mode Veille (01h-05h)", eco_metrics["unique_cells_optimized"], "Cellules"),
        ("Énergie Économisée par An", f"{eco_metrics['annual_kwh_saved']:,} kWh", "kWh / an"),
        ("Émissions CO2 Évitées par An", f"{eco_metrics['annual_co2_saved_tons']} Tonnes CO2", "t CO2 / an"),
        ("Équivalent Absorbé par la Forêt", f"~{eco_metrics['equivalent_trees_planted']:,} Arbres", "Arbres équivalents")
    ]

    for idx, (label, val, unit) in enumerate(rse_cards, start=12):
        ws1[f"A{idx}"] = label
        ws1[f"D{idx}"] = val
        ws1[f"A{idx}"].font = Font(bold=True)
        ws1[f"D{idx}"].font = Font(bold=True, color=COLOR_NAVY)
        ws1[f"A{idx}"].border = THIN_BORDER
        ws1[f"D{idx}"].border = THIN_BORDER

    # -------------------------------------------------------------------------
    # ONGLET 2 : ANALYSE DES PERFORMANCES RADIO
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Performance Radio")
    ws2.views.sheetView[0].showGridLines = True

    headers_perf = [
        "Site ID", "Nom du Site", "Cellule ID", "Techno", "Bande", 
        "RSRP Moyen (dBm)", "SINR Moyen (dB)", "PRB Max (%)", 
        "CDR Moyen (%)", "HOSR Moyen (%)", "Diagnostic Réseau"
    ]

    # En-tête
    ws2.append(headers_perf)
    for col_num, header in enumerate(headers_perf, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Remplissage des données
    for r_idx, row in perf_summary.iterrows():
        row_data = [
            row['site_id'], row['site_name'], row['cell_id'], row['technology'], row['frequency_band'],
            round(row['avg_rsrp'], 1), round(row['avg_sinr'], 1), round(row['max_prb_util'], 1),
            round(row['avg_cdr'], 2), round(row['avg_hosr'], 2), row['diagnostic']
        ]
        ws2.append(row_data)
        
        current_row = ws2.max_row
        is_degraded = row['is_degraded']
        bg_color = COLOR_RED_ALERT if is_degraded else (COLOR_BG_LIGHT if current_row % 2 == 0 else "FFFFFF")

        for col_num in range(1, len(headers_perf) + 1):
            c = ws2.cell(row=current_row, column=col_num)
            c.border = THIN_BORDER
            c.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            if col_num in [6, 7, 8, 9, 10]:
                c.alignment = Alignment(horizontal="right")

    # -------------------------------------------------------------------------
    # ONGLET 3 : PLAN D'ACTION & EXPORT ATOLL
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet(title="Plan d'Action ATOLL")
    ws3.views.sheetView[0].showGridLines = True

    headers_reco = [
        "Site ID", "Cellule ID", "Techno", "Bande", 
        "Tilt Elec Actuel", "Tilt Elec Proposé", "Azimuth", 
        "Action Recommandée", "Statut ATOLL", "Remarques Ingénierie"
    ]

    ws3.append(headers_reco)
    for col_num, header in enumerate(headers_reco, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_SKY_BLUE, end_color=COLOR_SKY_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in reco_summary.iterrows():
        row_data = [
            row['site_id'], row['cell_id'], row['technology'], row['frequency_band'],
            row['current_tilt_elec'], row['proposed_tilt_elec'], row['current_azimuth'],
            row['action_type'], row['atoll_status'], row['comment']
        ]
        ws3.append(row_data)

        current_row = ws3.max_row
        status = row['atoll_status']
        bg_color = COLOR_GREEN_OK if status == "Modified" else ("FFF2CC" if status == "Proposed" else "FFFFFF")

        for col_num in range(1, len(headers_reco) + 1):
            c = ws3.cell(row=current_row, column=col_num)
            c.border = THIN_BORDER
            c.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

    # Adjust Column Widths pour tous les onglets
    for sheet in [ws1, ws2, ws3]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    print(f"📊 Rapport Excel mis en forme et généré avec succès : {excel_path}\n")
    return excel_path


if __name__ == "__main__":
    from data_ingestion import fetch_merged_network_data
    from kpi_analyzer import analyze_radio_performance, calculate_decarbonization_impact
    from radio_dimensioning import generate_radio_recommendations

    try:
        data = fetch_merged_network_data()
        perf = analyze_radio_performance(data)
        eco = calculate_decarbonization_impact(data)
        reco = generate_radio_recommendations(perf)

        create_styled_excel_report(perf, reco, eco)
    except Exception as e:
        print(f"⚠️ Erreur lors de la génération du rapport Excel : {e}")